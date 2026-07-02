#!/usr/bin/env python3
"""
Auto-populate CS Reward Time Tracker from the weekly rota + database.

Usage:
    python3 /tmp/fill_tracker.py 2026-04-03

Argument is the Friday that starts the week (Fri-Thu).
Generates a new .xlsx, fills in roles from rota + actuals from DB, uploads to Google Drive.
"""
import sys
import os
import re
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

# ── Config ──────────────────────────────────────────────────────────────────
ROTA_SHEET_ID = '1CMSEZSb-4D4mO6iPb8tVSaAPsZT5KZst9VSXH4bpi0Y'
CREDS_PATH = Path.home() / '.config/juno/claude-code/google-credentials.json'
TRACKER_TEMPLATE = Path.home() / 'Downloads' / 'CS Reward Time Tracker.xlsx'
OUTPUT_DIR = Path.home() / 'Downloads'
DB_URL = os.environ['STAFF_APP_LOOKER_POSTGRES_URL']

# Tracker agents (in order they appear in the spreadsheet)
CORE_PHONES = ['Becky', 'Elida', 'Fionn', 'Kate']
# Harry (Courtney) joined the reward tracker 2026-06-11; Roseanne (Yasmin)
# from 2026-06-11. Charne kept in the list — her column is hidden on the
# live tracker.
WIDER_TEAM = ['Charne', 'Clare', 'Cris', 'Erika', 'Harriet', 'Harry',
              'Kirsty', 'Lizzie', 'Lucy', 'Noemi', 'Roseanne',
              'Sophie', 'Tara']
ALL_AGENTS = CORE_PHONES + WIDER_TEAM

# DB full name mapping (first name -> full name in database)
DB_NAMES = {
    'Becky': 'Becky Smith',
    'Elida': 'Elida Gizli',
    'Fionn': 'Fionn Burrows',
    'Kate': "Kate O'Neill",
    'Kirsty': 'Kirsty Rowley',
    'Charne': 'Charne Wilson',
    'Clare': 'Clare Brown',
    'Cris': 'Cris Macagi',
    'Erika': 'Erika Frolova',
    'Harriet': 'Harriet Clifton-Sprigg',
    'Lizzie': 'Lizzie Williamson',
    'Lucy': 'Lucy Riordan',
    'Noemi': 'Noemi Sip',
    'Harry': 'Harry McNicholas',
    'Roseanne': 'Roseanne Brooks-Brown',
    'Sophie': 'Sophie Maloney',
    'Tara': 'Tara Dunkley',
}
# Reverse: full name -> first name
FIRST_NAMES = {v: k for k, v in DB_NAMES.items()}

# Rota role -> tracker dropdown value
ROLE_MAP = {
    'Inbound phones': 'Phones',
    'Triage only': 'Triage',
    'Triage + lender chasing': 'Triage + Chasing',
    'Triage and Video Calls': 'Triage + Video Calls',
    'Case setup only': 'ICS',
    'Chasing': 'Call & Chase',
    'Team lead': 'Off',
    'Non working day': 'Off',
    'Annual leave': 'Holiday',
    'Unplanned absence': 'Sick',
    'Training': 'Training',
    'Case setup + inbound phones': 'ICS',
    'Case set up + lender chasing': 'ICS',
}

# Compound roles: "Inbound phones + X" -> primary is Phones
def map_rota_role(rota_val):
    """Map a rota role string to a tracker dropdown value."""
    if not rota_val:
        return ''
    rota_val = rota_val.strip()
    # Direct match
    if rota_val in ROLE_MAP:
        return ROLE_MAP[rota_val]
    # Compound: "Inbound phones + ..." -> Phones
    if rota_val.startswith('Inbound phones'):
        return 'Phones'
    # Compound: "Triage + ..." (not lender chasing)
    if rota_val.startswith('Triage'):
        return 'Triage'
    # Compound: "Case setup + ..."
    if rota_val.startswith('Case setup') or rota_val.startswith('Case set up'):
        return 'ICS'
    print(f"  WARNING: Unknown rota role '{rota_val}' - leaving blank")
    return ''

# Tracker column layout (from create_tracker_v4.py with Fri-first order)
# Days in tracker order: Fri, Mon, Tue, Wed, Thu
INFO_COLS = 3  # A=Week Ending, B=Agent Name, C=Primary Role
CPD = 6        # Columns per day: Role, Baseline, Stretch, Actual, BaselineMet, StretchMet
SS = 34        # Summary section start column
# Day blocks start at col D(4), each 6 wide
# Fri: D-I (cols 4-9), Mon: J-O (10-15), Tue: P-U (16-21), Wed: V-AA (22-27), Thu: AB-AG (28-33)

# Working Hours tab in tracker template — column for each day (1-indexed)
# B=Mon, C=Tue, D=Wed, E=Thu, F=Fri (matches template layout)
TEMPLATE_WH_DAY_COL = {0: 6, 1: 2, 2: 3, 3: 4, 4: 5}  # day_idx (Fri-first) -> col

# Standard targets (must match tracker template's Targets tab)
STANDARD_TARGETS = {
    'Phones':           {'baseline': 72,  'stretch': 82},
    'Triage':           {'baseline': 130, 'stretch': 180},
    'Call & Chase':     {'baseline': 95,  'stretch': 120},
    'ICS':              {'baseline': 275, 'stretch': 305},
    'Triage + Chasing': {'baseline': 104, 'stretch': 144},
    'Triage + Video Calls': {'baseline': 100, 'stretch': 140},
}


def _ensure_targets_rows(wb):
    """Ensure the Targets tab has a row for every STANDARD_TARGETS role.

    The tracker's per-role baseline/stretch cells are sheet formulas that
    VLOOKUP Targets!$A:$C (full columns), so a role missing from the Targets
    tab leaves its targets blank — and row position doesn't matter. We write
    missing roles into a blank row (the template's gap) or append at the
    bottom, and keep existing rows' values in sync. No row inserts (the
    footnote row is merged, which insert_rows mangles). Idempotent.
    """
    from openpyxl.cell.cell import MergedCell
    if 'Targets' not in wb.sheetnames:
        return
    ws = wb['Targets']

    existing, empty_rows = {}, []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None or (isinstance(a, str) and not a.strip()):
            empty_rows.append(r)
            continue
        if isinstance(a, str) and a.strip().lower().startswith('these are full-time'):
            continue  # footnote
        existing[str(a).strip()] = r

    def _set(r, c, v):
        cell = ws.cell(r, c)
        if isinstance(cell, MergedCell):
            return
        cell.value = v

    for role, t in STANDARD_TARGETS.items():
        if role in existing:
            _set(existing[role], 2, t['baseline'])
            _set(existing[role], 3, t['stretch'])
            continue
        target_row = empty_rows.pop(0) if empty_rows else (ws.max_row + 1)
        _set(target_row, 1, role)
        _set(target_row, 2, t['baseline'])
        _set(target_row, 3, t['stretch'])

# Cell colours
COLOUR_PRO_RATA = 'FFF2CC'   # Pale yellow
COLOUR_OVERRIDE = 'FCE5CD'   # Pale orange

# Daily Notes pro-rata trigger keywords (case-insensitive substring match in Note column)
DEFAULT_PRO_RATA_KEYWORDS = [
    'half day', 'annual leave', ' al ', 'al ', ' al', 'medical', 'dentist',
    'doctor', 'appointment', 'appt', 'early finish', 'late start', 'school',
    'blood test', 'hospital', 'amended hours', 'reward time',
]

# Recurring weekly pro-rata adjustments — applied every week regardless of Daily Notes.
# day_idx: 0=Fri, 1=Mon, 2=Tue, 3=Wed, 4=Thu (Fri-first tracker order)
RECURRING_WEEKLY_PRO_RATA = [
    {'agent': 'Noemi', 'day_idx': 3, 'hours_lost': 1.0, 'reason': 'Wednesday standup'},
]

# Rota Staff View — header row containing agent first names (0-indexed)
ROTA_HEADER_ROW = 4


def build_rota_agent_cols(header_row):
    """Map each tracked agent's first name to its column index in the rota header.
    Rebuilt on every run so column reorders in the rota don't silently
    point at the wrong person."""
    lookup = {}
    for idx, cell in enumerate(header_row):
        name = cell.strip()
        # First occurrence wins — names repeat in the "PHONES SECONDARIES"
        # block (cols ~34+) and we want the primary roster columns.
        if name in ALL_AGENTS and name not in lookup:
            lookup[name] = idx
    missing = [a for a in ALL_AGENTS if a not in lookup]
    if missing:
        raise RuntimeError(
            f"Rota header missing expected agents: {missing}. "
            f"Header row was: {header_row}"
        )
    return lookup


def parse_friday(date_str):
    """Parse YYYY-MM-DD and verify it's a Friday."""
    d = datetime.strptime(date_str, '%Y-%m-%d').date()
    if d.weekday() != 4:  # 4 = Friday
        print(f"WARNING: {date_str} is a {d.strftime('%A')}, not a Friday!")
    return d


def get_week_dates(friday):
    """Return the 5 dates in tracker order: Fri, Mon, Tue, Wed, Thu."""
    return [
        friday,                      # Fri (day 0)
        friday + timedelta(days=3),  # Mon (day 1)
        friday + timedelta(days=4),  # Tue (day 2)
        friday + timedelta(days=5),  # Wed (day 3)
        friday + timedelta(days=6),  # Thu (day 4)
    ]


def read_rota(friday):
    """Read the rota for the given week, return {agent: {day_idx: role}}."""
    from google.oauth2.credentials import Credentials
    import gspread

    print("Reading rota...")
    creds = Credentials.from_authorized_user_file(str(CREDS_PATH))
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(ROTA_SHEET_ID)
    # Roles display as emoji on Staff View; read the hidden "Roles (canonical)" mirror which
    # translates them back to the plain-text role names this script matches on. Same layout.
    ws = ss.worksheet("Roles (canonical)")
    all_data = ws.get_all_values()

    rota_agent_cols = build_rota_agent_cols(all_data[ROTA_HEADER_ROW])

    dates = get_week_dates(friday)
    # Build lookup: rota date string -> day_idx
    # Rota uses D/M/YY format (e.g., "3/4/26")
    date_to_idx = {}
    for idx, d in enumerate(dates):
        rota_fmt = f"{d.day}/{d.month}/{str(d.year)[-2:]}"
        date_to_idx[rota_fmt] = idx

    roles = {agent: {} for agent in ALL_AGENTS}
    found_dates = set()

    for row in all_data[5:]:  # Data starts after header rows
        date_str = row[1].strip()
        if date_str in date_to_idx:
            day_idx = date_to_idx[date_str]
            found_dates.add(date_str)
            for agent, col in rota_agent_cols.items():
                rota_val = row[col].strip() if col < len(row) else ''
                tracker_role = map_rota_role(rota_val)
                roles[agent][day_idx] = tracker_role

    missing = set(date_to_idx.keys()) - found_dates
    if missing:
        raise RuntimeError(
            f"Rota not ready: {len(missing)} of 5 dates missing ({missing}). "
            "Please update the rota before running the tracker."
        )
    print(f"  Found all 5 dates in rota")

    return roles


def read_working_hours():
    """Read the rota's Working Hours tab. Returns {agent: {day_idx: hours}}.

    day_idx uses tracker order (Fri=0, Mon=1, Tue=2, Wed=3, Thu=4).
    Falls back to None if tab is unreadable; caller should handle.
    """
    from google.oauth2.credentials import Credentials
    import gspread

    print("Reading rota Working Hours tab...")
    creds = Credentials.from_authorized_user_file(str(CREDS_PATH))
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(ROTA_SHEET_ID)
    try:
        ws = ss.worksheet("Working Hours")
    except Exception as e:
        print(f"  WARNING: Could not open Working Hours tab: {e}")
        return None

    rows = ws.get_all_values()
    if not rows:
        return None
    # Header: Agent, Mon Hrs, Tue Hrs, Wed Hrs, Thu Hrs, Fri Hrs, Weekly Total, Schedule
    # tracker day_idx → header col index (0-based)
    DAY_TO_COL = {1: 1, 2: 2, 3: 3, 4: 4, 0: 5}  # Mon, Tue, Wed, Thu, Fri

    hours = {}
    for row in rows[1:]:
        if not row or not row[0].strip():
            continue
        name = row[0].strip()
        if name not in ALL_AGENTS:
            continue
        per_day = {}
        for day_idx, col in DAY_TO_COL.items():
            val = row[col].strip().rstrip('.') if col < len(row) else ''
            try:
                per_day[day_idx] = float(val) if val else 0.0
            except ValueError:
                per_day[day_idx] = 0.0
        hours[name] = per_day
    print(f"  Read working hours for {len(hours)} agents")
    return hours


def read_overrides():
    """Read the rota's Overrides tab. Returns dict with two keys:
       'individual': list of {agent, role, baseline, stretch, reason}
       'standing':   {rule_name: bool active}
    """
    from google.oauth2.credentials import Credentials
    import gspread

    print("Reading rota Overrides tab...")
    creds = Credentials.from_authorized_user_file(str(CREDS_PATH))
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(ROTA_SHEET_ID)
    try:
        ws = ss.worksheet("Overrides")
    except Exception as e:
        print(f"  WARNING: No Overrides tab found ({e}) — using defaults")
        return {'individual': [], 'standing': {'within_1_of_stretch': True,
                                                'id_skips_below_baseline': True},
                'keywords': DEFAULT_PRO_RATA_KEYWORDS}

    rows = ws.get_all_values()
    individual = []
    standing = {}
    keywords = []
    section = None
    for row in rows:
        if not row or all(c.strip() == '' for c in row):
            continue
        first = row[0].strip()
        if first == 'INDIVIDUAL TARGET OVERRIDES':
            section = 'ind_header'
            continue
        if first == 'STANDING RULES (applied automatically by the script)':
            section = 'std_header'
            continue
        if first.startswith('DAILY NOTES PRO-RATA KEYWORDS'):
            section = 'kw_header'
            continue
        if section == 'ind_header' and first == 'Name':
            section = 'ind'
            continue
        if section == 'std_header' and first == 'Rule':
            section = 'std'
            continue
        if section == 'kw_header' and first == 'Keyword':
            section = 'kw'
            continue
        if section == 'ind':
            name = row[0].strip()
            role = row[1].strip() if len(row) > 1 else ''
            baseline_s = row[2].strip() if len(row) > 2 else ''
            stretch_s = row[3].strip() if len(row) > 3 else ''
            reason = row[4].strip() if len(row) > 4 else ''
            active = (row[5].strip().upper() if len(row) > 5 else 'Y')
            if not name or active != 'Y':
                continue
            try:
                baseline = int(baseline_s) if baseline_s else None
                stretch = int(stretch_s) if stretch_s else None
            except ValueError:
                print(f"  WARNING: Bad override values for {name}: baseline={baseline_s!r} stretch={stretch_s!r}")
                continue
            individual.append({'agent': name, 'role': role, 'baseline': baseline,
                               'stretch': stretch, 'reason': reason})
        elif section == 'std':
            rule = row[0].strip()
            active = (row[2].strip().upper() if len(row) > 2 else 'Y')
            if rule:
                standing[rule] = (active == 'Y')
        elif section == 'kw':
            kw = row[0].strip().lower()
            if kw:
                keywords.append(kw)

    if not keywords:
        keywords = DEFAULT_PRO_RATA_KEYWORDS
    if not standing:
        standing = {'within_1_of_stretch': True, 'id_skips_below_baseline': True}

    print(f"  Read {len(individual)} individual overrides, {sum(standing.values())} active standing rules, {len(keywords)} keywords")
    return {'individual': individual, 'standing': standing, 'keywords': keywords}


_TIME_RE = re.compile(r'^\s*(\d{1,2})(?:[:\.](\d{2}))?\s*(am|pm|AM|PM)?\s*$')


def _parse_time(s, fallback_pm_hint=False):
    """Parse a single clock time like '9', '9:30', '9am', '13:30'. Returns hour as float,
    or None if unparseable. fallback_pm_hint: if True and no am/pm marker, assume PM
    for hours 1..7 (so '5pm' from a '1-5pm' range works when only the second has 'pm')."""
    if s is None:
        return None
    s = s.strip().rstrip(' .')
    if not s:
        return None
    m = _TIME_RE.match(s)
    if not m:
        return None
    hh = int(m.group(1))
    mm = int(m.group(2)) if m.group(2) else 0
    ampm = m.group(3)
    if ampm:
        ampm = ampm.lower()
        if ampm == 'pm' and hh < 12:
            hh += 12
        elif ampm == 'am' and hh == 12:
            hh = 0
    elif fallback_pm_hint and 1 <= hh <= 7:
        hh += 12
    if hh > 24 or mm > 59:
        return None
    return hh + mm / 60.0


def parse_time_range(time_str):
    """Parse free-form time range like '1-5pm', '08:30-13:30', '3:30-5pm'.
    Returns hours_lost (float) or None if unparseable.
    """
    if not time_str:
        return None
    s = time_str.strip()
    # Normalise common typos: '::' → ':', '.' between digits → ':' (for HH.MM)
    s = s.replace('::', ':')
    # Split on '-' or '–' or 'to'
    parts = re.split(r'\s*[-–]\s*|\s+to\s+', s, maxsplit=1)
    if len(parts) != 2:
        return None
    left, right = parts
    # Determine if right side has am/pm; if so, propagate as hint to left
    right_has_ampm = bool(re.search(r'(am|pm)\s*$', right, re.IGNORECASE))
    end = _parse_time(right)
    if end is None:
        return None
    # For start, hint PM if right is PM and start has no marker and looks ambiguous
    start_has_ampm = bool(re.search(r'(am|pm)\s*$', left, re.IGNORECASE))
    hint_pm = right_has_ampm and not start_has_ampm and 'pm' in right.lower()
    start = _parse_time(left, fallback_pm_hint=hint_pm)
    if start is None:
        return None
    if end < start:
        # Crossed noon without explicit pm? Try adding 12 to end.
        if end + 12 > start and end + 12 <= 24:
            end += 12
        else:
            return None
    return round(end - start, 2)


def read_daily_notes(friday, dates):
    """Read Daily Notes for the 5 days. Returns list of entries:
       {date, day_idx, agent, time_str, role_str, note, hours_lost, parsed, reason_kw}
    """
    from google.oauth2.credentials import Credentials
    import gspread

    print("Reading Daily Notes...")
    creds = Credentials.from_authorized_user_file(str(CREDS_PATH))
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(ROTA_SHEET_ID)
    try:
        ws = ss.worksheet("Daily Notes")
    except Exception as e:
        print(f"  WARNING: Could not open Daily Notes tab: {e}")
        return []

    rows = ws.get_all_values()
    # Build date lookup as in read_rota
    date_to_idx = {}
    for idx, d in enumerate(dates):
        date_to_idx[f"{d.day}/{d.month}/{str(d.year)[-2:]}"] = (idx, d)

    entries = []
    # Skip header rows; header is at row index 3 (Date, Time, Name, Role, Note, ...)
    for row in rows[4:]:
        if not row or len(row) < 5:
            continue
        date_str = row[0].strip()
        if date_str not in date_to_idx:
            continue
        day_idx, day_date = date_to_idx[date_str]
        time_str = row[1].strip()
        agent = row[2].strip()
        role_str = row[3].strip()
        note = row[4].strip()
        if agent not in ALL_AGENTS:
            continue
        hours_lost = parse_time_range(time_str)
        entries.append({
            'date': day_date,
            'day_idx': day_idx,
            'agent': agent,
            'time_str': time_str,
            'role_str': role_str,
            'note': note,
            'hours_lost': hours_lost,
        })
    print(f"  Found {len(entries)} Daily Notes entries in window")
    return entries


def is_pro_rata_entry(note, role_str, keywords):
    """Decide whether a Daily Notes entry is an absence we should pro-rata."""
    # Annual leave / Unplanned absence in the role string always counts
    if role_str.lower() in ('annual leave', 'unplanned absence'):
        return True
    if not note:
        return False
    n = ' ' + note.lower() + ' '
    for kw in keywords:
        if kw in n:
            return True
    return False


def is_split_role_entry(note, role_str):
    """Detect ambiguous split-role mid-day switches that we should NOT auto-pro-rata."""
    n = note.lower()
    return ('split' in n and 'role' in n) or ('mid-day' in n) or ('switch' in n and 'role' in n)


def is_id_skip_entry(note):
    n = note.lower()
    return 'id skip' in n or 'id check delay' in n or 'id delay' in n


def find_pdt_table(cur):
    """Find the freshest pdt_things_done_by_person table in looker_scratch.

    Looker creates multiple PDT copies with opaque hash-based names — sorting
    by name doesn't reliably give the newest data.  Instead, check each table's
    MAX(completed_at) and return the one with the most recent data.
    """
    cur.execute("""
        SELECT table_schema || '.' || table_name
        FROM information_schema.tables
        WHERE table_name LIKE '%pdt_things_done_by_person'
    """)
    candidates = [r[0] for r in cur.fetchall()]
    if not candidates:
        raise RuntimeError("Could not find pdt_things_done_by_person table")

    best_table = None
    best_date = None
    for tbl in candidates:
        cur.execute(f"SELECT MAX(DATE(doable_or_enquiry_action_completed_at)) FROM {tbl}")
        max_date = cur.fetchone()[0]
        if max_date is not None and (best_date is None or max_date > best_date):
            best_date = max_date
            best_table = tbl

    if not best_table:
        raise RuntimeError("All pdt_things_done_by_person tables are empty")
    return best_table


def query_db(friday):
    """Query the database for actuals and skips. Returns (phone_data, things_data, skips_data)."""
    import psycopg2

    dates = get_week_dates(friday)
    start = min(dates)
    end = max(dates) + timedelta(days=1)  # Exclusive upper bound

    print(f"Querying database for {start} to {end}...")
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()

    # Find the latest PDT table
    pdt = find_pdt_table(cur)
    print(f"  Using PDT: {pdt}")

    # 1. Phone calls (inbound + outbound)
    print("  Querying phone_activity...")
    cur.execute("""
        SELECT DATE(started_at) as day, user_name, direction,
               COUNT(*) FILTER (WHERE answered_at IS NOT NULL
                   AND CASE WHEN direction = 'outbound' THEN duration > 25 ELSE true END) as answered
        FROM public.phone_activity
        WHERE started_at >= %s AND started_at < %s
          AND direction IN ('inbound', 'outbound')
          AND user_name IS NOT NULL
        GROUP BY 1, 2, 3
    """, (start, end))
    # {(first_name, date, direction): count}
    phone_data = defaultdict(int)
    for day, user_name, direction, count in cur.fetchall():
        first = FIRST_NAMES.get(user_name)
        if first:
            phone_data[(first, day, direction)] = count

    # 2. Things done (triage + ICS) - uses Looker PDT
    print("  Querying things done...")
    cur.execute(f"""
        SELECT DATE(doable_or_enquiry_action_completed_at) as day,
               staff_member_full_name, completed_method, COUNT(*) as cnt
        FROM {pdt}
        WHERE doable_or_enquiry_action_completed_at >= %s
          AND doable_or_enquiry_action_completed_at < %s
          AND staff_member_full_name IS NOT NULL
        GROUP BY 1, 2, 3
    """, (start, end))
    # {(first_name, date): {method: count}}
    things_data = defaultdict(lambda: defaultdict(int))
    for day, full_name, method, cnt in cur.fetchall():
        first = FIRST_NAMES.get(full_name)
        if first:
            things_data[(first, day)][method] = cnt

    # 3. Skips (weekly total) - "can't do" on work_allocation, joined via case_allocation for staff name
    print("  Querying skips...")
    cur.execute(f"""
        WITH staff_map AS (
            SELECT DISTINCT staff_member_id, staff_member_full_name
            FROM {pdt}
            WHERE staff_member_full_name IS NOT NULL
        )
        SELECT sm.staff_member_full_name, COUNT(DISTINCT wa.id) as skips
        FROM public.work_allocation wa
        JOIN public.case_allocation ca ON wa.case_allocation_id = ca.id
        JOIN staff_map sm ON ca.staff_member_id = sm.staff_member_id
        WHERE wa.db_completion_status = 'cant_do'
          AND wa.completed_at >= %s AND wa.completed_at < %s
        GROUP BY 1
    """, (start, end))
    skips_data = {}
    for full_name, skip_count in cur.fetchall():
        first = FIRST_NAMES.get(full_name)
        if first:
            skips_data[first] = skip_count

    conn.close()
    print("  DB queries complete")
    return phone_data, things_data, skips_data


def compute_actual(agent, day_date, role, phone_data, things_data):
    """Compute the actual number for an agent on a given day based on their role."""
    if role in ('Off', 'Holiday', 'Sick', 'Training', ''):
        return None

    # Don't write an actual for days that haven't happened yet (today or future).
    # Leaving the cell blank lets the TL View Reward formula show "⏳ Pending"
    # instead of a misleading "❌ None" from a day defaulting to 0 activity.
    from datetime import date as _date
    if day_date >= _date.today():
        return None

    if role == 'Phones':
        return phone_data.get((agent, day_date, 'inbound'), 0)

    if role == 'Call & Chase':
        return phone_data.get((agent, day_date, 'outbound'), 0)

    if role in ('Triage', 'Triage + Video Calls'):
        # Video calls add no separate countable metric — triage throughput
        # is emails archived for both.
        if (agent, day_date) not in things_data:
            return None  # PDT data not yet available for this day
        methods = things_data[(agent, day_date)]
        return methods.get('email archived', 0)

    if role == 'ICS':
        if (agent, day_date) not in things_data:
            return None  # PDT data not yet available for this day
        return sum(things_data[(agent, day_date)].values())

    if role == 'Triage + Chasing':
        # Combined: emails archived + connected outbound calls
        if (agent, day_date) not in things_data:
            return None  # PDT data not yet available — can't compute combined metric
        methods = things_data[(agent, day_date)]
        emails = methods.get('email archived', 0)
        outbound = phone_data.get((agent, day_date, 'outbound'), 0)
        return emails + outbound

    return None


def sync_working_hours(ws_wh, working_hours):
    """Update the tracker's Working Hours tab with values from rota.
    ws_wh: openpyxl worksheet for the 'Working Hours' tab.
    working_hours: {agent: {day_idx: hours}}
    Returns: {agent: row_number} mapping for downstream pro-rata writes.
    """
    agent_rows = {}
    for row in ws_wh.iter_rows(min_row=2, max_row=ws_wh.max_row):
        if not row[0].value:
            continue
        name = str(row[0].value).strip()
        if name in working_hours:
            agent_rows[name] = row[0].row
            for day_idx, col in TEMPLATE_WH_DAY_COL.items():
                hrs = working_hours[name].get(day_idx, 0.0)
                ws_wh.cell(row=row[0].row, column=col, value=hrs)
    return agent_rows


def apply_pro_rata(ws_wh, wh_agent_rows, daily_notes, working_hours, keywords, friday=None):
    """Reduce hours in the tracker's Working Hours tab for absence entries.

    Returns:
        applied: list of dicts (agent, day_idx, date, hours_lost, hours_before, hours_after, factor, reason)
        skipped: list of dicts for entries that triggered keyword match but couldn't be applied
    """
    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color=COLOUR_PRO_RATA, end_color=COLOUR_PRO_RATA, fill_type='solid')

    applied = []
    skipped = []

    # Aggregate hours lost per (agent, day) — multiple entries on the same day add up
    daily_totals = defaultdict(float)
    daily_reasons = defaultdict(list)
    for e in daily_notes:
        if not is_pro_rata_entry(e['note'], e['role_str'], keywords):
            continue
        if is_split_role_entry(e['note'], e['role_str']):
            skipped.append({**e, 'skip_reason': 'split-role'})
            continue
        if e['hours_lost'] is None:
            skipped.append({**e, 'skip_reason': 'unparseable_time'})
            continue
        if e['role_str'].lower() in ('annual leave', 'unplanned absence'):
            # Already handled by the script (role mapped to Holiday/Sick) — skip pro-rata
            continue
        daily_totals[(e['agent'], e['day_idx'])] += e['hours_lost']
        daily_reasons[(e['agent'], e['day_idx'])].append(e)

    # Recurring weekly adjustments (e.g. Noemi's Wednesday standup hour).
    # day_idx (Fri-first): Fri=0, Mon=1, Tue=2, Wed=3, Thu=4
    day_offset = {0: 0, 1: 3, 2: 4, 3: 5, 4: 6}
    for rule in RECURRING_WEEKLY_PRO_RATA:
        agent, day_idx = rule['agent'], rule['day_idx']
        date_str = ''
        if friday is not None:
            date_str = (friday + timedelta(days=day_offset[day_idx])).strftime('%Y-%m-%d')
        daily_totals[(agent, day_idx)] += rule['hours_lost']
        daily_reasons[(agent, day_idx)].append({
            'date': date_str,
            'note': rule['reason'],
            'role_str': '',
            'agent': agent,
            'day_idx': day_idx,
        })

    for (agent, day_idx), hours_lost in daily_totals.items():
        if agent not in wh_agent_rows:
            skipped.append({'agent': agent, 'day_idx': day_idx, 'skip_reason': 'agent_not_in_template'})
            continue
        row = wh_agent_rows[agent]
        col = TEMPLATE_WH_DAY_COL[day_idx]
        before = working_hours.get(agent, {}).get(day_idx, 0.0)
        after = max(0.0, round(before - hours_lost, 2))
        # Write reduced hours
        cell = ws_wh.cell(row=row, column=col, value=after)
        cell.fill = fill
        factor = round(after / before, 2) if before > 0 else 0
        applied.append({
            'agent': agent,
            'day_idx': day_idx,
            'date': daily_reasons[(agent, day_idx)][0]['date'],
            'hours_before': before,
            'hours_after': after,
            'hours_lost': hours_lost,
            'factor': factor,
            'reasons': [e['note'] for e in daily_reasons[(agent, day_idx)]],
        })

    return applied, skipped


def apply_individual_overrides(ws_data, agent_rows, roles, dates, working_hours, individual_overrides):
    """Replace baseline/stretch formulas with override-based formulas where applicable.

    The new formula: =ROUND(<override>*'Working Hours'!<day_col><wh_row>/8,0)
    This still pro-rates via Working Hours, so any subsequent day-level adjustments
    (e.g. half day) flow through automatically.
    """
    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color=COLOUR_OVERRIDE, end_color=COLOUR_OVERRIDE, fill_type='solid')

    applied = []

    # Build lookup of overrides by (agent, role)
    override_map = {}
    for ov in individual_overrides:
        override_map[(ov['agent'], ov['role'])] = ov

    # Working Hours columns in tracker template (B=Mon, C=Tue, D=Wed, E=Thu, F=Fri)
    wh_col_letter = {0: 'F', 1: 'B', 2: 'C', 3: 'D', 4: 'E'}

    for agent, row in agent_rows.items():
        for day_idx in range(5):
            role = roles.get(agent, {}).get(day_idx, '')
            if not role or role in ('Off', 'Holiday', 'Sick', 'Training'):
                continue
            ov = override_map.get((agent, role))
            if not ov:
                continue
            # Day block: role at col 4 + day_idx*6 (1-based); baseline = role+1, stretch = role+2
            base_col = INFO_COLS + 1 + day_idx * CPD  # role col
            wh_row_for_agent = None
            # Find the working-hours row for this agent (same name in Working Hours tab)
            # It's used in the formula: VLOOKUP via $B<row> = agent name, but we need
            # an explicit row reference for the per-day hours column.
            # Since the WH tab can have a different row order, use VLOOKUP for safety.
            wh_lookup = f"VLOOKUP($B{row},'Working Hours'!$A:$G,{TEMPLATE_WH_DAY_COL[day_idx]},FALSE)"
            if ov['baseline'] is not None:
                cell = ws_data.cell(row=row, column=base_col + 1)
                cell.value = (
                    f"=IFERROR(IF({wh_lookup}=0,\"\","
                    f"ROUND({ov['baseline']}*{wh_lookup}/8,0)),\"\")"
                )
                cell.fill = fill
            if ov['stretch'] is not None:
                cell = ws_data.cell(row=row, column=base_col + 2)
                cell.value = (
                    f"=IFERROR(IF({wh_lookup}=0,\"\","
                    f"ROUND({ov['stretch']}*{wh_lookup}/8,0)),\"\")"
                )
                cell.fill = fill
            applied.append({
                'agent': agent,
                'day_idx': day_idx,
                'role': role,
                'baseline': ov['baseline'],
                'stretch': ov['stretch'],
                'reason': ov['reason'],
            })
    return applied


def apply_standing_overrides(ws_data, agent_rows, roles, dates, working_hours,
                              individual_overrides, daily_notes, phone_data,
                              things_data, standing_rules):
    """Apply within-1-of-stretch and ID-skips-below-baseline overrides.

    These compare the computed Actual against the (pro-rated) baseline/stretch
    and write a literal "Yes" into BaselineMet/StretchMet, replacing the formula.
    """
    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color=COLOUR_OVERRIDE, end_color=COLOUR_OVERRIDE, fill_type='solid')
    applied = []

    # Map for individual baseline/stretch overrides (already used; need same lookup here)
    override_map = {(o['agent'], o['role']): o for o in individual_overrides}

    # ID-skip flag per (agent, day_idx) from Daily Notes
    id_skip_set = set()
    if standing_rules.get('id_skips_below_baseline'):
        for e in daily_notes:
            if is_id_skip_entry(e['note']):
                id_skip_set.add((e['agent'], e['day_idx']))

    for agent, row in agent_rows.items():
        for day_idx in range(5):
            role = roles.get(agent, {}).get(day_idx, '')
            if not role or role in ('Off', 'Holiday', 'Sick', 'Training'):
                continue
            day_date = dates[day_idx]
            actual = compute_actual(agent, day_date, role, phone_data, things_data)
            if actual is None:
                continue
            hrs = working_hours.get(agent, {}).get(day_idx, 0.0)
            if hrs <= 0:
                continue
            # Determine effective baseline/stretch (override or standard), then pro-rata
            std = STANDARD_TARGETS.get(role)
            if not std:
                continue
            ov = override_map.get((agent, role))
            base_full = ov['baseline'] if ov and ov.get('baseline') is not None else std['baseline']
            stretch_full = ov['stretch'] if ov and ov.get('stretch') is not None else std['stretch']
            baseline = round(base_full * hrs / 8)
            stretch = round(stretch_full * hrs / 8)

            base_col = INFO_COLS + 1 + day_idx * CPD  # role col
            bmet_col = base_col + 4
            smet_col = base_col + 5

            # within_1_of_stretch
            if standing_rules.get('within_1_of_stretch') and actual == stretch - 1:
                cell = ws_data.cell(row=row, column=smet_col, value='Yes')
                cell.fill = fill
                applied.append({
                    'agent': agent, 'day_idx': day_idx, 'rule': 'within_1_of_stretch',
                    'actual': actual, 'stretch': stretch,
                })

            # id_skips_below_baseline
            if (standing_rules.get('id_skips_below_baseline')
                    and (agent, day_idx) in id_skip_set
                    and actual < baseline):
                cell = ws_data.cell(row=row, column=bmet_col, value='Yes')
                cell.fill = fill
                applied.append({
                    'agent': agent, 'day_idx': day_idx, 'rule': 'id_skips_below_baseline',
                    'actual': actual, 'baseline': baseline,
                })

    return applied


def validate(applied_pro_rata, skipped_pro_rata, applied_individual, applied_standing,
             daily_notes, keywords):
    """Return list of warning dicts: {level, message}.
    level in {'info', 'warn', 'error'}.
    """
    warnings = []

    # Suspicious pro-rata factors
    for e in applied_pro_rata:
        if e['factor'] < 0.25 and e['hours_after'] > 0:
            warnings.append({
                'level': 'warn',
                'message': (f"Suspicious pro-rata factor {e['factor']:.2f} for "
                            f"{e['agent']} on {e['date']} — only {e['hours_after']}h of "
                            f"{e['hours_before']}h. Verify Daily Notes entry."),
            })

    # Skipped entries that triggered keyword match
    for e in skipped_pro_rata:
        reason = e.get('skip_reason', 'unknown')
        if reason == 'unparseable_time':
            warnings.append({
                'level': 'warn',
                'message': (f"Could not parse time '{e['time_str']}' for {e['agent']} on "
                            f"{e['date']} ({e['note']!r}) — manual review needed."),
            })
        elif reason == 'split-role':
            warnings.append({
                'level': 'warn',
                'message': (f"Split-role day for {e['agent']} on {e['date']} — "
                            f"target not adjusted, Jo to review manually."),
            })
        elif reason == 'agent_not_in_template':
            warnings.append({
                'level': 'warn',
                'message': f"Agent {e['agent']} not in tracker template — pro-rata not applied.",
            })

    return warnings


def print_report(friday, applied_pro_rata, skipped_pro_rata, applied_individual,
                 applied_standing, warnings):
    """Print a structured report for the prompt to grab."""
    DAYS_LABEL = ['Fri', 'Mon', 'Tue', 'Wed', 'Thu']
    print()
    print("=" * 60)
    print("REPORT")
    print("=" * 60)
    print()
    print("PRO-RATA APPLIED")
    if applied_pro_rata:
        for e in sorted(applied_pro_rata, key=lambda x: (x['day_idx'], x['agent'])):
            day = DAYS_LABEL[e['day_idx']]
            reasons = '; '.join(e['reasons'])[:80]
            print(f"  {e['agent']} - {day} ({e['date']}): worked {e['hours_after']}h "
                  f"of {e['hours_before']}h (factor {e['factor']:.2f}) — {reasons}")
    else:
        print("  (none)")

    print()
    print("INDIVIDUAL TARGET OVERRIDES APPLIED")
    if applied_individual:
        # Group by agent/role for compact output
        by_agent = defaultdict(list)
        for e in applied_individual:
            by_agent[(e['agent'], e['role'])].append(e)
        for (agent, role), items in by_agent.items():
            ex = items[0]
            parts = []
            if ex['baseline'] is not None:
                parts.append(f"baseline={ex['baseline']}")
            if ex['stretch'] is not None:
                parts.append(f"stretch={ex['stretch']}")
            print(f"  {agent} - {role} ({len(items)} day(s)): {', '.join(parts)} — {ex['reason']}")
    else:
        print("  (none)")

    print()
    print("STANDING OVERRIDES APPLIED")
    if applied_standing:
        for e in sorted(applied_standing, key=lambda x: (x['day_idx'], x['agent'])):
            day = DAYS_LABEL[e['day_idx']]
            if e['rule'] == 'within_1_of_stretch':
                print(f"  {e['agent']} - {day}: within 1 of stretch "
                      f"(actual {e['actual']}, stretch {e['stretch']}) — StretchMet=Yes")
            elif e['rule'] == 'id_skips_below_baseline':
                print(f"  {e['agent']} - {day}: ID skips below baseline "
                      f"(actual {e['actual']}, baseline {e['baseline']}) — BaselineMet=Yes")
    else:
        print("  (none)")

    print()
    print("WARNINGS")
    if warnings:
        for w in warnings:
            print(f"  [{w['level'].upper()}] {w['message']}")
    else:
        print("  (none)")
    print("=" * 60)


def generate_and_fill(friday, roles, phone_data, things_data, skips_data,
                       working_hours=None, daily_notes=None, overrides=None):
    """Generate a fresh tracker .xlsx and fill in the data.

    If working_hours / daily_notes / overrides are provided, applies pro-rata
    and target overrides during the fill. Returns (output_path, report_dict).
    """
    from openpyxl import load_workbook
    import shutil

    dates = get_week_dates(friday)
    week_str = friday.strftime('%Y-%m-%d')
    output_path = OUTPUT_DIR / f'CS Reward Time Tracker - wc {week_str}.xlsx'

    # Step 1: Copy template
    print("Copying tracker template...")
    shutil.copy2(TRACKER_TEMPLATE, output_path)

    # Step 2: Open and fill
    print(f"Filling data into {output_path.name}...")
    wb = load_workbook(output_path)
    ws = wb['Data']

    # Agent rows in the tracker
    CORE_START = 4  # Row 4 is first core phones agent
    WIDER_START = CORE_START + len(CORE_PHONES) + 1  # +1 for section header

    agent_rows = {}
    for i, name in enumerate(CORE_PHONES):
        agent_rows[name] = CORE_START + i
    for i, name in enumerate(WIDER_TEAM):
        agent_rows[name] = WIDER_START + i

    for agent in ALL_AGENTS:
        row = agent_rows[agent]

        # Week ending date (Thursday = friday + 6 days)
        ws.cell(row=row, column=1, value=friday + timedelta(days=6))

        for day_idx in range(5):
            day_date = dates[day_idx]
            role = roles.get(agent, {}).get(day_idx, '')

            # Day block columns
            bc = INFO_COLS + 1 + day_idx * CPD  # Role column for this day

            # Write role
            if role:
                ws.cell(row=row, column=bc, value=role)

            # Write actual
            actual = compute_actual(agent, day_date, role, phone_data, things_data)
            if actual is not None:
                ws.cell(row=row, column=bc + 3, value=actual)  # Actual is 4th col in day block

        # Weekly skips
        skip_count = skips_data.get(agent, 0)
        ws.cell(row=row, column=SS + 6, value=skip_count)  # Weekly Skips column

        # Email stats for archive ratio (AZ=52, BA=53)
        emails_archived_total = 0
        emails_escalated_total = 0
        for day_idx in range(5):
            day_date = dates[day_idx]
            methods = things_data.get((agent, day_date), {})
            emails_archived_total += methods.get('email archived', 0)
            emails_escalated_total += methods.get('email escalated', 0)
        ws.cell(row=row, column=52, value=emails_archived_total)   # AZ: Emails Archived
        ws.cell(row=row, column=53, value=emails_escalated_total)  # BA: Emails Escalated

    # Step 3: Sync Working Hours from rota, apply pro-rata + overrides
    report = {'pro_rata': [], 'pro_rata_skipped': [], 'individual': [],
              'standing': [], 'warnings': []}
    if working_hours and 'Working Hours' in wb.sheetnames:
        ws_wh = wb['Working Hours']
        wh_rows = sync_working_hours(ws_wh, working_hours)

        # Apply pro-rata adjustments to Working Hours cells (formulas auto-recalc)
        if daily_notes is not None:
            keywords = (overrides or {}).get('keywords', DEFAULT_PRO_RATA_KEYWORDS)
            applied_pr, skipped_pr = apply_pro_rata(
                ws_wh, wh_rows, daily_notes, working_hours, keywords, friday=friday)
            report['pro_rata'] = applied_pr
            report['pro_rata_skipped'] = skipped_pr

            # Recompute working_hours dict to reflect post-pro-rata values
            # (so standing-override comparisons use the adjusted hours)
            wh_after = {a: dict(d) for a, d in working_hours.items()}
            for e in applied_pr:
                wh_after.setdefault(e['agent'], {})[e['day_idx']] = e['hours_after']
        else:
            wh_after = working_hours

        # Individual target overrides (from the Overrides tab)
        if overrides and overrides.get('individual'):
            applied_ind = apply_individual_overrides(
                ws, agent_rows, roles, dates, wh_after, overrides['individual'])
            report['individual'] = applied_ind

        # Standing overrides (within-1-of-stretch, ID skips)
        if overrides and overrides.get('standing'):
            applied_std = apply_standing_overrides(
                ws, agent_rows, roles, dates, wh_after,
                overrides.get('individual', []), daily_notes or [],
                phone_data, things_data, overrides['standing'])
            report['standing'] = applied_std

        # Validation
        report['warnings'] = validate(
            report['pro_rata'], report['pro_rata_skipped'],
            report['individual'], report['standing'],
            daily_notes or [], (overrides or {}).get('keywords', DEFAULT_PRO_RATA_KEYWORDS))

    # Keep the Targets tab in sync with STANDARD_TARGETS so the per-role
    # baseline/stretch VLOOKUPs resolve (e.g. Triage + Video Calls).
    _ensure_targets_rows(wb)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    return output_path, report


def upload_to_drive(file_path):
    """Upload the filled tracker to Google Drive as a new Google Sheet."""
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    print("Uploading to Google Drive...")
    creds = Credentials.from_authorized_user_file(str(CREDS_PATH))
    service = build('drive', 'v3', credentials=creds)

    media = MediaFileUpload(
        str(file_path),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    result = service.files().create(
        body={
            'name': file_path.stem,
            'mimeType': 'application/vnd.google-apps.spreadsheet',
        },
        media_body=media,
        fields='id',
    ).execute()

    file_id = result['id']
    url = f"https://docs.google.com/spreadsheets/d/{file_id}"
    print(f"Uploaded: {url}")
    return url


def last_friday():
    """Return last Friday's date (7 days ago if today is Friday)."""
    from datetime import date
    today = date.today()
    days_since_friday = (today.weekday() - 4) % 7
    if days_since_friday == 0:
        days_since_friday = 7  # If today is Friday, go back to LAST Friday
    return today - timedelta(days=days_since_friday)


def main():
    if len(sys.argv) == 2:
        friday = parse_friday(sys.argv[1])
    else:
        friday = last_friday()
        print(f"Auto-detected last Friday: {friday}")

    print(f"Week: {friday.strftime('%A %d %B %Y')} to {(friday + timedelta(days=6)).strftime('%A %d %B %Y')}")

    # Step 1: Read rota
    roles = read_rota(friday)
    for agent in ALL_AGENTS:
        day_roles = roles.get(agent, {})
        print(f"  {agent:10s}: {', '.join(day_roles.get(i, '-') for i in range(5))}")

    # Step 1b: Read Working Hours, Overrides, Daily Notes from rota
    dates = get_week_dates(friday)
    working_hours = read_working_hours()
    overrides = read_overrides()
    daily_notes = read_daily_notes(friday, dates)

    # Step 2: Query DB
    phone_data, things_data, skips_data = query_db(friday)

    # Step 3: Generate and fill (with pro-rata + overrides)
    output_path, report = generate_and_fill(
        friday, roles, phone_data, things_data, skips_data,
        working_hours=working_hours, daily_notes=daily_notes, overrides=overrides)

    # Step 4: Upload
    url = upload_to_drive(output_path)

    # Step 5: Add TL View + TL Calculator tabs (all tabs visible — this is Jo's full version)
    spreadsheet_id = url.split('/d/')[1].split('/')[0] if '/d/' in url else url.rsplit('/', 1)[-1]
    from setup_tl_view import get_sheets_service, create_tl_view, create_tl_calculator
    print("Setting up TL View and TL Calculator tabs...")
    sheets_service = get_sheets_service()
    create_tl_view(sheets_service, spreadsheet_id)
    create_tl_calculator(sheets_service, spreadsheet_id)
    print("  TL View + TL Calculator ready")

    # Step 6: Print structured report (pro-rata, overrides, warnings)
    print_report(friday, report['pro_rata'], report['pro_rata_skipped'],
                 report['individual'], report['standing'], report['warnings'])

    print(f"\nDone! Google Sheet: {url}")
    print(f"\nTo correct after rota updates, run:")
    print(f"  python3 /Users/joannejeffries/Documents/GitHub/cs-scripts/recompute_tracker.py "
          f"{spreadsheet_id} {friday.strftime('%Y-%m-%d')}")
    print(f"\nTo create the TL-friendly copy (tabs hidden), run:")
    print(f"  python3 -c \"from setup_tl_view import make_tl_copy, get_sheets_service; "
          f"print(make_tl_copy(get_sheets_service(), '{spreadsheet_id}'))\"")
    return url


if __name__ == '__main__':
    main()
